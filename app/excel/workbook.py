"""Workbook utilities and safe open/save operations.

Provides a lightweight wrapper around openpyxl.Workbook for reading and writing
while preserving styles, formulas, merged cells, column widths and row heights.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Tuple
import shutil
import time

from loguru import logger
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import coordinate_to_tuple, get_column_letter, column_index_from_string


@dataclass(slots=True)
class ExcelBlock:
    name: str
    anchor_cell: str
    row: int
    column: int


class WorkbookHandler:
    def __init__(self, path: str, data_only: bool = False) -> None:
        self.path = Path(path)
        self.data_only = data_only
        self.wb = None
        self.ws: Optional[Worksheet] = None

    def open(self, sheet_name: str) -> None:
        if not self.path.exists():
            logger.error("Workbook does not exist: {}", self.path)
            raise FileNotFoundError(f"Workbook not found: {self.path}")

        logger.info("Opening workbook {} (data_only={})", self.path, self.data_only)
        self.wb = load_workbook(self.path, data_only=self.data_only)
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Worksheet {sheet_name} not found in workbook")
        self.ws = self.wb[sheet_name]
        logger.info("Opened worksheet {sheet}", sheet=sheet_name)

    def save(self, backup: bool = True) -> None:
        if backup:
            self._create_backup()
        if self.wb is None:
            raise RuntimeError("Workbook not opened")
        self.wb.save(self.path)
        logger.info("Workbook saved: {}", self.path)

    def _create_backup(self) -> Path:
        bak_dir = Path("backups")
        bak_dir.mkdir(parents=True, exist_ok=True)
        ts = int(time.time())
        dst = bak_dir / f"{self.path.stem}_{ts}{self.path.suffix}"
        shutil.copy2(self.path, dst)
        logger.info("Backup created: {}", dst)
        return dst

    def get_cell(self, coord: str):
        if self.ws is None:
            raise RuntimeError("Worksheet not opened")
        return self.ws[coord]

    def get_cell_value(self, coord: str):
        cell = self.get_cell(coord)
        # If cell is part of merged cell range, return top-left value
        for merged in self.ws.merged_cells.ranges:
            if coord in merged:
                min_col, min_row, _, _ = merged.bounds
                top_left = f"{get_column_letter(min_col)}{min_row}"
                return self.ws[top_left].value
        return cell.value

    @staticmethod
    def coord_to_tuple(coord: str) -> Tuple[int, int]:
        col, row = coordinate_to_tuple(coord)
        return row, column_index_from_string(col)

    @staticmethod
    def tuple_to_coord(row: int, col: int) -> str:
        return f"{get_column_letter(col)}{row}"
